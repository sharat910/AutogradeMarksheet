from django.shortcuts import render_to_response,render
from django.http import HttpResponseRedirect,Http404,HttpResponse
import json
import openpyxl as o
from openpyxl.utils import get_column_letter, column_index_from_string,coordinate_from_string
from collections import defaultdict
import matplotlib.pyplot as plt
from .forms import *
from django.template import RequestContext
from random import randint

# Create your views here.
def FileUpload(request):
    form = UploadForm(request.POST or None,request.FILES or None)
    if form.is_valid():
        m = Marksheet(f=request.FILES['f'],n =request.POST['n'],t =request.POST['t'],c = request.POST['c'])
        m.save()
        return HttpResponseRedirect("/home/"+ str(m.id))
    context = {'form': form}
    return render_to_response("upload.html",context,context_instance = RequestContext(request))

def Home(request,mid):
    form = TestForm(request.POST or None)
    if form.is_valid():
        n = request.POST['bins']
        s = request.POST['binint']
        try:
            freeze = request.POST['freeze']
        except:
            freeze = False
        bins = []
        if n:
            k = 100/int(n)
            for i in range(0,100,k):
                bins.append(i)

            if not (k*int(n)) == 100:
                bins[-1] = 100
            else:
                bins += [100]
        elif s:
            a = s.split(',')
            arr = map(int,a)
            bins = [0]
            bins +=arr
            bins.append(100)
        else:
            raise Http404
        print bins

        m = Marksheet.objects.get(pk = mid)
        wb = o.load_workbook(filename = m.f.url[1:],data_only=True)
        #sheet = wb.get_sheet_by_name('Sheet1')
        sheet = wb.active
        marks = []
        xy = coordinate_from_string(m.c) # returns ('A',4)
        print xy
        col = column_index_from_string(xy[0]) # returns 1
        row = xy[1]
        for i in range(row,row + m.n):
            if sheet.cell(row = i,column = col).value:
                val = int(sheet.cell(row = i,column = col).value)
                marks.append(val)
            else:
                marks.append(0)
        print "Before3 Plot"
        print marks
        if not freeze:
            plt.hist(marks,bins,histtype= 'bar',rwidth='0.9',alpha = 0.7)
            plt.xlabel('mark')
            plt.ylabel('frequency')
            r = randint(10000,99999)
            plt.savefig('media/images/' +  str(mid) + '.png')
            plt.hold(False)
            return HttpResponseRedirect('/media/images/' +  str(mid) + '.png')
        else:
            grade = []
            if not len(bins) == 10:
                return render_to_response("error.html",{'mid':mid },context_instance = RequestContext(request))
            for mark in marks:
                if mark <= bins[1]:
                    grade.append('NC')
                elif mark <= bins[2]:
                    grade.append('E')
                elif mark <= bins[3]:
                    grade.append('D')
                elif mark <= bins[4]:
                    grade.append('C-')
                elif mark <= bins[5]:
                    grade.append('C')
                elif mark <= bins[6]:
                    grade.append('B-')
                elif mark <= bins[7]:
                    grade.append('B')
                elif mark <= bins[8]:
                    grade.append('A-')
                elif mark <= bins[9]:
                    grade.append('A')
            print len(marks),len(grade)
            sheet.cell(row=row-1,column=col+1,value= "Grades")
            for i,g in enumerate(grade):
                sheet.cell(row = row+i,column = col + 1, value = g)
            wb.save("media/files/graded"+str(mid) + ".xlsx")
            return HttpResponseRedirect('/media/files/graded'+str(mid) +'.xlsx')

    context = {'form': form}
    return render_to_response("home.html",context,context_instance = RequestContext(request))
